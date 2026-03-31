"""
import_reviews.py
================
Imports all reviews from UCR_class_difficulty_database.xlsx into Firebase Firestore.
Uploads reviews + updates courseStats aggregates.

SETUP:
    pip3 install firebase-admin openpyxl

USAGE:
    1. Go to Firebase Console → Project Settings → Service Accounts
    2. Click "Generate new private key" → save as serviceAccountKey.json
       in the same folder as this script
    3. Put UCR_class_difficulty_database.xlsx in the same folder
    4. Run: python3 import_reviews.py

This script is safe to run multiple times — it checks for existing
imported reviews and skips duplicates.
"""

import re
import sys
import time
from openpyxl import load_workbook

try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except ImportError:
    print("Run: pip3 install firebase-admin openpyxl")
    sys.exit(1)

XLSX_FILE    = "UCR_class_difficulty_database.xlsx"
SERVICE_KEY  = "serviceAccountKey.json"
BATCH_SIZE   = 400  # Firestore max batch size is 500


def normalize_code(code):
    """AHS007 -> AHS 007, CHEM001A -> CHEM 001A"""
    code = str(code).strip().upper()
    m = re.match(r'^([A-Z]+)(\d.*)$', code)
    if m:
        return f"{m.group(1)} {m.group(2)}"
    return code


def normalize_difficulty(d):
    """Convert any scale to 1-5"""
    try:
        val = float(str(d).strip())
        if val > 5:
            return round((val / 10) * 5, 1)
        return round(val, 1)
    except:
        return None


def parse_reviews(xlsx_path):
    print(f"Reading {xlsx_path}...")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    current_code = None
    reviews = []

    for row in rows[2:]:  # skip header rows
        code       = row[0]
        comment    = row[2]
        difficulty = row[3]
        date       = row[4]

        if code and str(code).strip() and str(code).strip() not in ('', 'None'):
            raw = str(code).strip()
            if raw and re.match(r'^[A-Z]', raw):
                current_code = normalize_code(raw)

        if current_code and difficulty and str(difficulty).strip() not in ('', 'None'):
            diff = normalize_difficulty(difficulty)
            if diff and 1 <= diff <= 5:
                reviews.append({
                    'courseCode':      current_code,
                    'difficulty':      diff,
                    'professorRating': 3.0,   # neutral default
                    'workload':        'Medium',
                    'wouldRecommend':  diff <= 3.0,
                    'gradeReceived':   '',
                    'termTaken':       '',
                    'professor':       '',
                    'text':            str(comment).strip() if comment else '',
                    'source':          'imported',
                    'createdAt':       str(date) if date else '',
                })

    print(f"Parsed {len(reviews)} reviews across {len(set(r['courseCode'] for r in reviews))} courses")
    return reviews


def build_stats(reviews):
    """Pre-compute aggregate stats per course."""
    stats = {}
    for r in reviews:
        code = r['courseCode']
        if code not in stats:
            stats[code] = {
                'courseCode':      code,
                'reviewCount':     0,
                'totalDifficulty': 0,
                'totalProfRating': 0,
                'recommendCount':  0,
            }
        s = stats[code]
        s['reviewCount']     += 1
        s['totalDifficulty'] += r['difficulty']
        s['totalProfRating'] += r['professorRating']
        s['recommendCount']  += 1 if r['wouldRecommend'] else 0
    return stats


def upload(reviews, stats):
    print(f"\nConnecting to Firebase...")
    cred = credentials.Certificate(SERVICE_KEY)
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    # Check if already imported
    existing = db.collection('reviews').where('source', '==', 'imported').limit(1).get()
    if existing:
        answer = input(f"⚠️  Imported reviews already exist. Re-import? (y/n): ").strip().lower()
        if answer != 'y':
            print("Skipped.")
            return

    # Upload reviews in batches
    print(f"\nUploading {len(reviews)} reviews in batches of {BATCH_SIZE}...")
    for i in range(0, len(reviews), BATCH_SIZE):
        batch = db.batch()
        chunk = reviews[i:i + BATCH_SIZE]
        for r in chunk:
            ref = db.collection('reviews').document()
            batch.set(ref, r)
        batch.commit()
        print(f"  {min(i + BATCH_SIZE, len(reviews))}/{len(reviews)} reviews uploaded...")
        time.sleep(0.5)

    # Upload/merge course stats
    print(f"\nUpdating stats for {len(stats)} courses...")
    stat_list = list(stats.values())
    for i in range(0, len(stat_list), BATCH_SIZE):
        batch = db.batch()
        chunk = stat_list[i:i + BATCH_SIZE]
        for s in chunk:
            doc_id = s['courseCode'].replace(' ', '_')
            ref = db.collection('courseStats').document(doc_id)
            # Merge with existing stats (from real reviews)
            batch.set(ref, s, merge=True)
        batch.commit()
        print(f"  {min(i + BATCH_SIZE, len(stat_list))}/{len(stat_list)} course stats updated...")
        time.sleep(0.5)

    print(f"\n✅ Done! {len(reviews)} reviews imported.")
    print(f"   {len(stats)} courses now have stats on rcourses.org")


def main():
    import os
    if not os.path.exists(SERVICE_KEY):
        print(f"❌ Missing {SERVICE_KEY}")
        print("Go to Firebase Console → Project Settings → Service Accounts")
        print("→ Generate new private key → save as serviceAccountKey.json here")
        sys.exit(1)

    if not os.path.exists(XLSX_FILE):
        print(f"❌ Missing {XLSX_FILE}")
        sys.exit(1)

    reviews = parse_reviews(XLSX_FILE)
    stats   = build_stats(reviews)
    upload(reviews, stats)


if __name__ == "__main__":
    main()

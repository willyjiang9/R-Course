"""
xlsx_to_js.py
=============
Converts ucr_courses.xlsx (from the scraper) into
src/data/allCourses.json for the React app.

USAGE:
    pip3 install openpyxl
    python3 xlsx_to_js.py

    (run from the ucr-courses/ project root)
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("Run: pip3 install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────
XLSX_FILE   = "ucr_all_terms.xlsx"   # path to your scraped xlsx
OUTPUT_FILE = "src/data/allCourses.json"
SHEET_NAME  = "Spring 2026"        # use the most current term sheet
                                    # or set to None to use first sheet
# ─────────────────────────────────────────────────────────────────

def load_courses(xlsx_path, sheet_name=None):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "Spring 2026" in wb.sheetnames:
        ws = wb["Spring 2026"]
    elif "All Courses" in wb.sheetnames:
        ws = wb["All Courses"]
    else:
        ws = wb.active

    print(f"Reading sheet: {ws.title}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Sheet is empty.")
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    print(f"Columns: {headers}")

    # Map column names to indices
    col = {h: i for i, h in enumerate(headers)}

    def get(row, name, default=""):
        i = col.get(name)
        if i is None: return default
        v = row[i]
        return str(v).strip() if v is not None else default

    courses = []
    seen = set()

    for row in rows[1:]:
        subject = get(row, "Subject Code")
        number  = get(row, "Course Number")
        if not subject or not number:
            continue

        full_code = get(row, "Full Code") or f"{subject} {number}"
        if full_code in seen:
            continue
        seen.add(full_code)

        # Skip graduate courses (number >= 200)
        try:
            num_int = int(''.join(filter(str.isdigit, number)))
            if num_int >= 200:
                continue
        except ValueError:
            pass

        courses.append({
            "fullCode":     full_code,
            "subject":      subject,
            "number":       number,
            "title":        get(row, "Title"),
            "units":        get(row, "Units"),
            "description":  get(row, "Description"),
            "prerequisites": get(row, "Prerequisites"),
            "scheduleType": get(row, "Schedule Type"),
        })

    return courses


def main():
    if not os.path.exists(XLSX_FILE):
        print(f"❌ File not found: {XLSX_FILE}")
        print("Make sure ucr_courses.xlsx is in the same folder as this script.")
        sys.exit(1)

    print(f"Loading {XLSX_FILE}...")
    courses = load_courses(XLSX_FILE, SHEET_NAME)
    print(f"Found {len(courses)} unique undergraduate courses")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(courses, f, indent=2)

    print(f"✅ Written to {OUTPUT_FILE}")
    print()
    print("Next step — update src/App.jsx line 13:")
    print('  // Remove this line:')
    print('  import { SAMPLE_COURSES, ... } from \'./data/courses.js\'')
    print('  const COURSES = SAMPLE_COURSES')
    print()
    print('  // Add this line instead:')
    print('  import COURSES from \'./data/allCourses.json\'')


if __name__ == "__main__":
    main()
